"""
Plan / Timeline export (FN-PLAN-05, PL-7/PL-8).

Excel via openpyxl (Thai is fine as UTF-8). PDF via reportlab with the bundled
Sarabun font embedded (PRD §9: never rely on the PDF lib's default font for
Thai). A lightweight Gantt is drawn below the table from start/end dates.
"""
from pathlib import Path

from django.http import HttpResponse

FONT_DIR = Path(__file__).resolve().parent / "fonts"
_FONTS_REGISTERED = False


def _f(value, default=0.0):
    return float(value) if value is not None else default


# --- Excel -----------------------------------------------------------------
def plan_to_xlsx(project) -> HttpResponse:
    import openpyxl
    from openpyxl.styles import Font

    items = list(project.plan_items.all())
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plan"
    ws.append([f"แผนงาน: {project.project_name} ({project.project_code or '-'})"])
    ws["A1"].font = Font(bold=True, size=14)

    header = ["เฟส", "งาน", "วันเริ่ม", "วันสิ้นสุด", "Manday", "Milestone"]
    ws.append(header)
    for c in ws[2]:
        c.font = Font(bold=True)

    total = 0.0
    for it in items:
        md = _f(it.manday)
        total += md
        ws.append([
            it.phase,
            it.task,
            it.start_date.isoformat() if it.start_date else "",
            it.end_date.isoformat() if it.end_date else "",
            md,
            "✓" if it.is_milestone else "",
        ])
    ws.append([])
    ws.append(["", "", "", "รวม manday", total, ""])
    ws[ws.max_row][3].font = Font(bold=True)
    ws[ws.max_row][4].font = Font(bold=True)

    for col, width in zip("ABCDEF", (26, 32, 14, 14, 10, 12)):
        ws.column_dimensions[col].width = width

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="plan_{project.id}.xlsx"'
    wb.save(resp)
    return resp


# --- PDF -------------------------------------------------------------------
def _register_fonts():
    global _FONTS_REGISTERED
    if _FONTS_REGISTERED:
        return
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    pdfmetrics.registerFont(TTFont("Sarabun", str(FONT_DIR / "Sarabun-Regular.ttf")))
    pdfmetrics.registerFont(TTFont("Sarabun-Bold", str(FONT_DIR / "Sarabun-Bold.ttf")))
    _FONTS_REGISTERED = True


_TH_MONTH = ["ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.",
             "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]


def _truncate(text, limit):
    text = text or ""
    return text if len(text) <= limit else text[: limit - 1] + "…"


def _gantt_flowable(items, width, group_by_phase=False):
    """Gantt with month axis, track + green progress fill, milestone diamonds
    and a today line — mirrors the on-screen chart. With group_by_phase, each
    phase gets a bold header row with a slim summary bar spanning its range."""
    from datetime import date as _date

    from reportlab.graphics.shapes import Drawing, Line, Polygon, Rect, String
    from reportlab.lib.colors import HexColor

    from .services import plan_item_progress

    dated = [i for i in items if i.start_date and i.end_date]
    if not dated:
        return None

    # Row list: ("phase", label, start_ord, end_ord) | ("item", plan_item)
    rows = []
    if group_by_phase:
        by_phase = []
        for it in dated:
            if not by_phase or by_phase[-1][0] != it.phase:
                by_phase.append((it.phase, []))
            by_phase[-1][1].append(it)
        for phase, phase_items in by_phase:
            rows.append((
                "phase",
                phase,
                min(i.start_date for i in phase_items).toordinal(),
                max(i.end_date for i in phase_items).toordinal(),
            ))
            rows.extend(("item", i) for i in phase_items)
    else:
        rows = [("item", i) for i in dated]

    min_d = min(i.start_date for i in dated).toordinal() - 2
    max_d = max(i.end_date for i in dated).toordinal() + 2
    span = max(max_d - min_d, 1)
    row_h, bar_h, head_h, label_w = 18, 10, 16, 170
    chart_w = width - label_w - 10

    def X(o):
        return label_w + (o - min_d) / span * chart_w

    height = head_h + len(rows) * row_h + 4
    d = Drawing(width, height)

    track = HexColor("#cbd5e1")
    phase_c = HexColor("#94a3b8")
    prog_c = HexColor("#059669")
    mile_c = HexColor("#2563eb")
    grid = HexColor("#e2e8f0")
    today_c = HexColor("#e11d48")
    txt = HexColor("#475569")
    strong = HexColor("#1e293b")
    body_top = height - head_h

    # month gridlines + labels
    cur = _date(_date.fromordinal(min_d).year, _date.fromordinal(min_d).month, 1)
    while cur.toordinal() <= max_d:
        x = X(cur.toordinal())
        d.add(Line(x, 0, x, body_top, strokeColor=grid))
        d.add(String(x + 2, body_top + 4,
                     f"{_TH_MONTH[cur.month - 1]} {str(cur.year + 543)[2:]}",
                     fontName="Sarabun", fontSize=7, fillColor=txt))
        cur = _date(cur.year + 1, 1, 1) if cur.month == 12 else _date(cur.year, cur.month + 1, 1)

    # rows (top to bottom)
    for idx, row in enumerate(rows):
        y = body_top - (idx + 1) * row_h + (row_h - bar_h) / 2
        if row[0] == "phase":
            _, label, start_o, end_o = row
            d.add(String(0, y + 1, _truncate(label, 30), fontName="Sarabun-Bold", fontSize=7.5, fillColor=strong))
            x1 = X(start_o)
            bw = max(X(end_o + 1) - x1, 2)
            d.add(Rect(x1, y + bar_h / 2 - 1.5, bw, 3, fillColor=phase_c, strokeColor=None))
            continue
        it = row[1]
        indent = 8 if group_by_phase else 0
        d.add(String(indent, y + 1, _truncate(it.task, 30), fontName="Sarabun", fontSize=7, fillColor=txt))
        x1 = X(it.start_date.toordinal())
        bw = max(X(it.end_date.toordinal() + 1) - x1, 2)
        d.add(Rect(x1, y, bw, bar_h, fillColor=track, strokeColor=None))
        pr = plan_item_progress(it)
        if pr:
            d.add(Rect(x1, y, max(bw * pr, 1), bar_h, fillColor=prog_c, strokeColor=None))
        if it.is_milestone:
            cx, cy = x1 + bw, y + bar_h / 2
            d.add(Polygon([cx, cy + 5, cx + 5, cy, cx, cy - 5, cx - 5, cy],
                          fillColor=mile_c, strokeColor=None))

    # today line
    t = _date.today().toordinal()
    if min_d <= t <= max_d:
        d.add(Line(X(t), 0, X(t), body_top, strokeColor=today_c, strokeWidth=1))

    d.add(Line(label_w, 0, label_w, body_top, strokeColor=grid))
    return d


def plan_to_pdf(project) -> HttpResponse:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    _register_fonts()
    items = list(project.plan_items.all())

    resp = HttpResponse(content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="plan_{project.id}.pdf"'

    # Explicit 30pt margins — the gantt drawing is sized to page-60, so the
    # default 72pt margins would clip it at the right edge.
    doc = SimpleDocTemplate(
        resp, pagesize=landscape(A4), title="ASTRO Plan",
        leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30,
    )
    styles = getSampleStyleSheet()
    h = ParagraphStyle("h", parent=styles["Title"], fontName="Sarabun-Bold", fontSize=16)
    cell = ParagraphStyle("c", fontName="Sarabun", fontSize=9, leading=12)

    story = [Paragraph(f"แผนงาน: {project.project_name} ({project.project_code or '-'})", h), Spacer(1, 6)]

    header = ["เฟส", "งาน", "วันเริ่ม", "วันสิ้นสุด", "Manday", "MS"]
    rows = [[Paragraph(t, ParagraphStyle("hh", fontName="Sarabun-Bold", fontSize=9)) for t in header]]
    total = 0.0
    for it in items:
        md = _f(it.manday)
        total += md
        rows.append([
            Paragraph(it.phase or "", cell),
            Paragraph(it.task or "", cell),
            Paragraph(it.start_date.isoformat() if it.start_date else "-", cell),
            Paragraph(it.end_date.isoformat() if it.end_date else "-", cell),
            Paragraph(f"{md:g}", cell),
            Paragraph("◆" if it.is_milestone else "", cell),
        ])
    rows.append([Paragraph("", cell), Paragraph("", cell), Paragraph("", cell),
                 Paragraph("รวม", ParagraphStyle("b", fontName="Sarabun-Bold", fontSize=9)),
                 Paragraph(f"{total:g}", ParagraphStyle("b", fontName="Sarabun-Bold", fontSize=9)),
                 Paragraph("", cell)])

    table = Table(rows, colWidths=[55 * mm, 90 * mm, 28 * mm, 28 * mm, 22 * mm, 12 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (4, 0), (5, -1), "CENTER"),
    ]))
    story.append(table)
    story.append(Spacer(1, 14))

    gantt = _gantt_flowable(items, width=landscape(A4)[0] - 60)
    if gantt is not None:
        story.append(Paragraph("Gantt", ParagraphStyle("g", fontName="Sarabun-Bold", fontSize=12)))
        story.append(Spacer(1, 4))
        story.append(gantt)

    doc.build(story)
    return resp


def export_plan(project, fmt: str) -> HttpResponse:
    if fmt == "pdf":
        return plan_to_pdf(project)
    return plan_to_xlsx(project)


# --- Client-facing progress report (FN: ส่งลูกค้าเป็นเอกสารอัพเดตความคืบหน้า) --
def _th_date(value):
    if not value:
        return "-"
    return f"{value.day} {_TH_MONTH[value.month - 1]} {value.year + 543}"


def progress_report_pdf(project) -> HttpResponse:
    """
    รายงานความคืบหน้าโครงการสำหรับส่งลูกค้า: หัวเอกสาร (โครงการ/ลูกค้า/PO/
    ระยะเวลา/% รวม/สถานะ) + ตารางความคืบหน้ารายเฟส + Gantt แบบจัดกลุ่มเฟส.
    Confidential fields (value_thb) are deliberately absent.
    """
    import datetime

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    from apps.projects.models import HealthStatus

    from .services import plan_item_progress, project_progress

    _register_fonts()
    items = list(project.plan_items.all())

    resp = HttpResponse(content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="progress_report_{project.id}.pdf"'

    # 30pt margins to match the gantt drawing width (page-60) — default 72pt
    # margins clip the drawing's right edge.
    doc = SimpleDocTemplate(
        resp, pagesize=landscape(A4), title="รายงานความคืบหน้าโครงการ",
        leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30,
    )
    title = ParagraphStyle("t", fontName="Sarabun-Bold", fontSize=17, leading=22)
    label = ParagraphStyle("l", fontName="Sarabun-Bold", fontSize=9.5, leading=13)
    value = ParagraphStyle("v", fontName="Sarabun", fontSize=9.5, leading=13)
    section = ParagraphStyle("s", fontName="Sarabun-Bold", fontSize=12, leading=16)

    progress = project_progress(project)
    progress_txt = f"{round(progress * 100)}%" if progress is not None else "ยังไม่มีข้อมูล"
    health_label = HealthStatus(project.health_status).label
    status_txt = f"{health_label}" + (f" — {project.health_reason}" if project.health_reason else "")

    story = [
        Paragraph("รายงานความคืบหน้าโครงการ", title),
        Spacer(1, 8),
    ]

    meta_rows = [
        ["โครงการ", f"{project.project_name} ({project.project_code or '-'})",
         "ระยะเวลาโครงการ", f"{_th_date(project.start_date)} – {_th_date(project.end_date)}"],
        ["ลูกค้า", project.client.client_name,
         "วันที่ออกเอกสาร", _th_date(datetime.date.today())],
        ["ผู้รับผิดชอบโครงการ", (project.po_user.full_name or project.po_user.email) if project.po_user else "-",
         "ความคืบหน้ารวม", progress_txt],
        ["สถานะ", status_txt, "", ""],
    ]
    meta = Table(
        [[Paragraph(a, label), Paragraph(b, value), Paragraph(c, label), Paragraph(d_, value)]
         for a, b, c, d_ in meta_rows],
        colWidths=[35 * mm, 100 * mm, 35 * mm, 65 * mm],
    )
    meta.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f1f5f9")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#f1f5f9")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("SPAN", (1, 3), (3, 3)),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.extend([meta, Spacer(1, 14)])

    # --- per-phase progress table -----------------------------------------
    if items:
        phases = []
        for it in items:  # model order (sort_order)
            if not phases or phases[-1][0] != it.phase:
                phases.append((it.phase, []))
            phases[-1][1].append(it)

        head = ParagraphStyle("hh", fontName="Sarabun-Bold", fontSize=9.5)
        rows = [[Paragraph(t, head) for t in ["เฟส", "ช่วงวันที่", "Manday รวม", "ความคืบหน้า"]]]
        for phase, phase_items in phases:
            starts = [i.start_date for i in phase_items if i.start_date]
            ends = [i.end_date for i in phase_items if i.end_date]
            period = f"{_th_date(min(starts) if starts else None)} – {_th_date(max(ends) if ends else None)}"
            manday = sum(_f(i.manday) for i in phase_items)
            num = den = 0.0
            for i in phase_items:
                pr = plan_item_progress(i)
                if pr is None or i.manday is None:
                    continue
                num += pr * _f(i.manday)
                den += _f(i.manday)
            pct = f"{round(num / den * 100)}%" if den > 0 else "ยังไม่มีข้อมูล"
            rows.append([
                Paragraph(phase or "-", value),
                Paragraph(period, value),
                Paragraph(f"{manday:g}", value),
                Paragraph(pct, value),
            ])

        table = Table(rows, colWidths=[95 * mm, 80 * mm, 30 * mm, 30 * mm])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (2, 0), (3, -1), "CENTER"),
        ]))
        story.extend([Paragraph("ความคืบหน้ารายเฟส", section), Spacer(1, 4), table, Spacer(1, 14)])

    gantt = _gantt_flowable(items, width=landscape(A4)[0] - 60, group_by_phase=True)
    if gantt is not None:
        story.extend([Paragraph("แผนงาน (Gantt)", section), Spacer(1, 4), gantt])

    doc.build(story)
    return resp
